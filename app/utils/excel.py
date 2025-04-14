import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models.student import Student, Class
from app.models.user import User, Role
from app import db
from datetime import datetime

def create_student_import_template():
    """创建学生导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息导入模板"
    
    # 定义表头
    headers = [
        '学号*', '姓名*', '性别*', '生日', '班级ID*', '邮箱*', '密码', 
        '电话', '状态', '地址', '简介'
    ]
    
    # 设置列宽
    for i in range(len(headers)):
        ws.column_dimensions[chr(65 + i)].width = 15
    
    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 添加示例数据
    example_data = [
        '2024001', '张三', 'male', '2000-01-01', '1', 'zhangsan@example.com',
        '123456', '13800138000', 'active', '北京市海淀区', '计算机专业学生'
    ]
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.alignment = Alignment(horizontal='center')
    
    return wb

def process_student_import(file_path, skip_header=True, duplicate_handling='update', auto_create_class=False):
    """处理学生数据导入
    
    Args:
        file_path: Excel文件路径
        skip_header: 是否跳过表头
        duplicate_handling: 重复数据处理方式 ('update'/'skip')
        auto_create_class: 是否自动创建不存在的班级
        
    Returns:
        tuple: (成功导入数量, 错误记录列表)
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path)
        
        # 跳过表头行
        if skip_header:
            df = df.iloc[1:]
        
        success_count = 0
        error_records = []
        
        # 获取学生角色ID
        student_role = Role.query.filter_by(name='Student').first()
        if not student_role:
            raise ValueError('Student role not found')
        
        # 处理每一行数据
        for index, row in df.iterrows():
            try:
                # 验证必填字段
                if pd.isna(row['学号*']) or pd.isna(row['姓名*']) or pd.isna(row['性别*']) or \
                   pd.isna(row['班级ID*']) or pd.isna(row['邮箱*']):
                    error_records.append({
                        'row': index + 2,
                        'error': '必填字段不能为空'
                    })
                    continue
                
                # 验证班级ID是否存在
                class_id = int(row['班级ID*'])
                class_obj = Class.query.get(class_id)
                
                # 如果班级不存在且启用了自动创建选项
                if not class_obj:
                    if auto_create_class:
                        from app.models.teacher import Department
                        # 获取第一个部门作为默认部门
                        default_department = Department.query.first()
                        
                        # 创建一个基础班级
                        current_year = datetime.now().year
                        class_obj = Class(
                            id=class_id,
                            name=f"自动创建班级-{class_id}",
                            code=f"AUTO-{class_id}",
                            grade=str(current_year),  # 使用当前年份作为年级
                            major_id=1,  # 使用默认专业ID
                            description=f"通过学生导入自动创建的班级，创建于{datetime.now().strftime('%Y-%m-%d')}",
                            created_at=datetime.now(),
                            updated_at=datetime.now()
                        )
                        db.session.add(class_obj)
                        db.session.flush()
                    else:
                        error_records.append({
                            'row': index + 2,
                            'error': f'班级ID {class_id} 不存在'
                        })
                        continue
                
                # 检查学号是否重复
                existing_student = Student.query.filter_by(student_id=str(row['学号*'])).first()
                if existing_student:
                    if duplicate_handling == 'skip':
                        error_records.append({
                            'row': index + 2,
                            'error': f'学号 {row["学号*"]} 已存在，已跳过'
                        })
                        continue
                    # 如果是更新模式，删除旧记录
                    db.session.delete(existing_student)
                    if existing_student.user:
                        db.session.delete(existing_student.user)
                
                # 创建用户账号
                user = User(
                    username=str(row['学号*']),
                    email=str(row['邮箱*']),
                    name=str(row['姓名*']),
                    phone=str(row['电话']) if not pd.isna(row['电话']) else None,
                    password=str(row['密码']) if not pd.isna(row['密码']) else '123456',  # 默认密码
                    role_id=student_role.id,
                    is_active=True
                )
                db.session.add(user)
                db.session.flush()  # 获取用户ID
                
                # 创建学生记录
                student = Student(
                    student_id=str(row['学号*']),
                    name=str(row['姓名*']),
                    gender=str(row['性别*']),
                    birthday=pd.to_datetime(row['生日']).date() if not pd.isna(row['生日']) else None,
                    class_id=class_id,
                    phone=str(row['电话']) if not pd.isna(row['电话']) else None,
                    status=str(row['状态']) if not pd.isna(row['状态']) else 'active',
                    address=str(row['地址']) if not pd.isna(row['地址']) else None,
                    user_id=user.id
                )
                db.session.add(student)
                success_count += 1
                
            except Exception as e:
                error_records.append({
                    'row': index + 2,
                    'error': str(e)
                })
                db.session.rollback()
                continue
        
        # 提交所有成功的记录
        if success_count > 0:
            db.session.commit()
        
        return success_count, error_records
        
    except Exception as e:
        return 0, [{'row': 0, 'error': f'文件处理失败：{str(e)}'}] 